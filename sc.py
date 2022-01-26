from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
import pandas as pd
from selenium.webdriver.common.by import By
from pandas.plotting import table
import matplotlib.pyplot as plt

wb = load_workbook('input_sheet.xlsx')
sheets = wb.sheetnames
sh1 = wb['Need to Upload Confirmation Rec']

ls = []
for i in range(2, 12):
    data = sh1.cell(i, 1).value


    s = Service("C:\\Users\\Yashp\\Downloads\\Compressed\\Selenium\\chromedriver.exe")
    driver = webdriver.Chrome(service=s)
    driver.get("https://brevets-patents.ic.gc.ca/opic-cipo/cpd/eng/search/number.html")
    search_box = driver.find_element(By.XPATH, '//*[@id="query"]').send_keys(data)
    search_btn = driver.find_element(By.XPATH, '//*[@id="button-group"]/div/input[1]').click()

    sample = driver.current_url
    scraper = pd.read_html(
        "https://brevets-patents.ic.gc.ca/opic-cipo/cpd/eng/patent/737283/summary.html?type=number_search&tabs1Index=tabs1_1")

    table_data = scraper[1]
    table_data.to_excel(f'CA_{data}.xlsx')

    ax = plt.subplot(120, frame_on=False)
    ax.xaxis.set_visible(0)
    ax.yaxis.set_visible(0)
    table(ax, table_data, loc='upper center')
    plt.savefig(f'CA_{data}.pdf')
    driver.quit()

